from __future__ import annotations

import csv
from collections import Counter
from dataclasses import dataclass
from io import BytesIO, StringIO
import re
from typing import Any, Literal
import unicodedata
from zipfile import BadZipFile

from fastapi import HTTPException, status
from openpyxl import load_workbook
from sqlalchemy import func, or_, select
from sqlalchemy.orm import Session

from app.models.imported_employee import ImportedEmployee
from app.schemas.imported_employee import EmployeeImportOptions

DetectedFormat = Literal["csv", "xlsx"]
RowSeverity = Literal["warning", "error"]
RowStatus = Literal["importable", "incomplete", "error"]

PREVIEW_ROW_LIMIT = 8
EMAIL_REGEX = re.compile(r"^[^\s@]+@[^\s@]+\.[^\s@]+$")

FIELD_ORDER = (
    "full_name",
    "email",
    "department_name",
    "job_profile",
    "employee_identifier",
    "employee_code",
    "status",
)
EDITABLE_FIELDS = {
    "full_name",
    "email",
    "department_name",
    "job_profile",
    "employee_identifier",
    "employee_code",
    "status",
}
FIELD_LABELS: dict[str, str] = {
    "full_name": "Nom",
    "email": "Email",
    "department_name": "Departement",
    "job_profile": "Fonction",
    "employee_identifier": "Identifiant",
    "employee_code": "Matricule",
    "status": "Statut",
}
FIELD_ALIASES: dict[str, set[str]] = {
    "full_name": {
        "nom complet",
        "collaborateur",
        "employee",
        "employee name",
        "employee full name",
        "employe",
        "full name",
        "full_name",
        "name",
        "nom",
    },
    "email": {
        "adresse email",
        "courriel",
        "e mail",
        "e-mail",
        "email",
        "email address",
        "mail",
    },
    "employee_identifier": {
        "employee id",
        "employee identifier",
        "identifiant",
        "identifiant collaborateur",
        "identifiant employee",
        "identifiant employe",
        "identifiant rh",
        "login",
        "user id",
        "username",
    },
    "department_name": {
        "business unit",
        "department",
        "departement",
        "direction",
        "equipe",
        "service",
    },
    "job_profile": {
        "employee profile",
        "fonction",
        "job profile",
        "position",
        "poste",
        "profil",
        "role",
        "title",
    },
    "status": {
        "etat",
        "situation",
        "state",
        "statut",
        "status",
    },
    "employee_code": {
        "code employe",
        "employee code",
        "employee number",
        "matricule",
        "matricule rh",
        "numero employe",
        "staff id",
    },
}


@dataclass(slots=True)
class RowIssue:
    code: str
    severity: RowSeverity
    message: str
    field_name: str | None = None
    fixable: bool = False


@dataclass(slots=True)
class ParsedEmployeeRow:
    row_number: int
    full_name: str | None
    email: str | None
    employee_identifier: str | None
    employee_code: str | None
    department_name: str | None
    job_profile: str | None
    status: str
    identity_key: str | None
    row_status: RowStatus
    issues: list[RowIssue]
    duplicate_reason: str | None = None


@dataclass(slots=True)
class EmployeeFileAnalysis:
    file_name: str
    detected_format: DetectedFormat
    total_rows: int
    valid_rows: int
    ready_rows: int
    incomplete_rows: int
    invalid_rows: int
    duplicate_rows: int
    error_rows: int
    quality_score: int
    anomalies_count: int
    fixable_anomalies: int
    global_notice: str | None
    recognized_columns: list[dict[str, str]]
    available_columns: list[str]
    field_mappings: list[dict[str, Any]]
    missing_required_fields: list[str]
    warnings: list[str]
    suggestions: list[dict[str, Any]]
    preview_rows: list[dict[str, Any]]
    importable_rows: list[ParsedEmployeeRow]


def _ensure_imported_employee_storage(db: Session) -> None:
    ImportedEmployee.__table__.create(bind=db.get_bind(), checkfirst=True)


def _clean_text(value: Any) -> str | None:
    if value is None:
        return None
    cleaned = " ".join(str(value).replace("\u00a0", " ").split()).strip()
    return cleaned or None


def _normalize_header(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value).encode("ascii", "ignore").decode("ascii")
    normalized = normalized.lower().strip()
    normalized = re.sub(r"[^a-z0-9]+", " ", normalized)
    return " ".join(normalized.split())


def _normalize_search(value: str | None) -> str | None:
    cleaned = _clean_text(value)
    return cleaned.lower() if cleaned else None


def _normalize_identifier(value: str | None) -> str | None:
    cleaned = _clean_text(value)
    return cleaned.lower() if cleaned else None


def _normalize_email(value: str | None) -> str | None:
    cleaned = _normalize_search(value)
    if not cleaned:
        return None
    return cleaned if EMAIL_REGEX.match(cleaned) else None


def _normalize_status(value: str | None) -> str:
    normalized = _normalize_search(value)
    if normalized in {"inactive", "inactif", "desactive", "disabled", "offboarded"}:
        return "inactive"
    if normalized in {"suspended", "suspendu", "suspendue", "blocked", "bloque"}:
        return "suspended"
    return "active"


def _smart_title_case(value: str | None) -> str | None:
    cleaned = _clean_text(value)
    if not cleaned:
        return None
    if len(cleaned) <= 4 and cleaned.isupper():
        return cleaned
    if cleaned == cleaned.lower() or cleaned == cleaned.upper():
        parts = re.split(r"(\s+|[-'/])", cleaned.lower())
        return "".join(part.capitalize() if part.isalpha() else part for part in parts)
    return cleaned


def _auto_fix_email(value: str | None) -> str | None:
    cleaned = _clean_text(value)
    if not cleaned:
        return None

    compact = cleaned.replace(" ", "").lower()
    compact = compact.replace("(at)", "@").replace("[at]", "@")

    if compact.count("@") == 1:
        local_part, domain_part = compact.split("@", 1)
        if "." not in domain_part:
            domain_part = domain_part.replace(",", ".").replace(";", ".")
        compact = f"{local_part}@{domain_part}"

    return compact


def _build_identity_key(
    email: str | None,
    employee_identifier: str | None,
    employee_code: str | None,
) -> str | None:
    if email:
        return f"email:{email}"
    if employee_identifier:
        return f"identifier:{employee_identifier}"
    if employee_code:
        return f"code:{employee_code}"
    return None


def _column_match_score(header: str, field_name: str) -> int:
    normalized_header = _normalize_header(header)
    aliases = FIELD_ALIASES[field_name]
    if normalized_header in aliases:
        return 100

    header_tokens = set(normalized_header.split())
    best_score = 0

    for alias in aliases:
        alias_tokens = set(alias.split())
        overlap = len(header_tokens & alias_tokens)
        if overlap > 0:
            best_score = max(best_score, (overlap * 20) - abs(len(header_tokens) - len(alias_tokens)))
        if alias in normalized_header or normalized_header in alias:
            best_score = max(best_score, 12)

    return best_score


def _suggest_columns_for_field(headers: list[str], field_name: str, selected_header: str | None) -> list[str]:
    ranked_headers = sorted(
        headers,
        key=lambda header: (
            header != selected_header,
            -_column_match_score(header, field_name),
            header.lower(),
        ),
    )
    suggestions = [
        header
        for header in ranked_headers
        if header == selected_header or _column_match_score(header, field_name) > 0
    ]
    return suggestions[:4]


def _build_field_mappings(
    headers: list[str],
    options: EmployeeImportOptions,
) -> tuple[dict[str, str], list[dict[str, Any]], list[dict[str, str]], list[str], list[str]]:
    auto_detected: dict[str, str] = {}
    for header in headers:
        normalized_header = _normalize_header(header)
        for field_name, aliases in FIELD_ALIASES.items():
            if field_name in auto_detected:
                continue
            if normalized_header in aliases:
                auto_detected[field_name] = header

    manual_overrides = {
        field_name: source_column
        for field_name, source_column in options.mapping_overrides.items()
        if field_name in FIELD_ALIASES
    }

    field_to_header: dict[str, str] = {}
    field_mappings: list[dict[str, Any]] = []
    recognized_columns: list[dict[str, str]] = []

    for field_name in FIELD_ORDER:
        is_manually_assigned = field_name in manual_overrides
        overridden_header = manual_overrides.get(field_name)
        source_column = (
            overridden_header
            if overridden_header in headers
            else None
            if is_manually_assigned
            else auto_detected.get(field_name)
        )
        if source_column:
            field_to_header[field_name] = source_column
            recognized_columns.append(
                {
                    "field_name": field_name,
                    "source_column": source_column,
                }
            )

        helper_text = None
        if source_column is None and field_name in {"full_name", "email", "department_name", "job_profile"}:
            helper_text = "Associez cette colonne pour ameliorer la qualite des donnees."
        elif is_manually_assigned:
            helper_text = "Association manuelle active."

        field_mappings.append(
            {
                "field_name": field_name,
                "label": FIELD_LABELS[field_name],
                "source_column": source_column,
                "required": field_name in {"full_name", "email"},
                "confidence": "manual" if is_manually_assigned else ("high" if source_column else "none"),
                "manually_assigned": is_manually_assigned,
                "suggested_columns": _suggest_columns_for_field(headers, field_name, source_column),
                "helper_text": helper_text,
            }
        )

    missing_required_fields: list[str] = []
    if "full_name" not in field_to_header:
        missing_required_fields.append("Nom complet")
    if not any(field_name in field_to_header for field_name in ("email", "employee_identifier", "employee_code")):
        missing_required_fields.append("Email / identifiant / matricule")

    warnings: list[str] = []
    if "status" not in field_to_header:
        warnings.append("Aucune colonne statut reconnue: le statut 'actif' sera applique par defaut.")
    if "department_name" not in field_to_header:
        warnings.append(
            "Aucune colonne departement reconnue: un rattachement manuel pourra etre necessaire."
        )
    if "job_profile" not in field_to_header:
        warnings.append(
            "Aucune colonne fonction/profil reconnue: le champ restera a completer si besoin."
        )

    duplicate_mappings = [
        header
        for header, count in Counter(field_to_header.values()).items()
        if count > 1
    ]
    for duplicate_header in duplicate_mappings:
        warnings.append(
            f"La colonne '{duplicate_header}' est associee a plusieurs champs. Verifiez le mapping."
        )

    return field_to_header, field_mappings, recognized_columns, missing_required_fields, warnings


def _read_csv_rows(content: bytes) -> tuple[list[str], list[dict[str, str]]]:
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Le fichier CSV doit etre encode en UTF-8.",
        ) from exc

    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel

    reader = csv.DictReader(StringIO(text), dialect=dialect)
    headers = [header.strip() for header in reader.fieldnames or [] if header and header.strip()]
    rows = [
        {key.strip(): (value or "") for key, value in row.items() if key is not None}
        for row in reader
        if any(_clean_text(value) for value in row.values())
    ]
    return headers, rows


def _read_xlsx_rows(content: bytes) -> tuple[list[str], list[dict[str, str]]]:
    try:
        workbook = load_workbook(filename=BytesIO(content), read_only=True, data_only=True)
    except BadZipFile as exc:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Le fichier Excel est invalide ou corrompu.",
        ) from exc
    except Exception as exc:  # pragma: no cover - defensive parsing guard
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Impossible de lire le fichier Excel fourni.",
        ) from exc

    worksheet = workbook.active
    raw_rows = worksheet.iter_rows(values_only=True)
    header_row = next(raw_rows, None)
    if header_row is None:
        return [], []

    headers = [
        _clean_text(value) or f"colonne_{index + 1}"
        for index, value in enumerate(header_row)
    ]
    rows: list[dict[str, str]] = []

    for row in raw_rows:
        row_dict = {
            headers[index]: "" if value is None else str(value)
            for index, value in enumerate(row[: len(headers)])
        }
        if any(_clean_text(value) for value in row_dict.values()):
            rows.append(row_dict)

    workbook.close()
    return headers, rows


def _read_uploaded_rows(file_name: str, content: bytes) -> tuple[DetectedFormat, list[str], list[dict[str, str]]]:
    normalized_name = file_name.lower().strip()
    if normalized_name.endswith(".csv"):
        headers, rows = _read_csv_rows(content)
        return "csv", headers, rows
    if normalized_name.endswith(".xlsx"):
        headers, rows = _read_xlsx_rows(content)
        return "xlsx", headers, rows

    raise HTTPException(
        status_code=status.HTTP_400_BAD_REQUEST,
        detail="Format non supporte. Utilisez un fichier CSV ou XLSX.",
    )


def _serialize_issue(issue: RowIssue) -> dict[str, Any]:
    return {
        "code": issue.code,
        "severity": issue.severity,
        "message": issue.message,
        "field_name": issue.field_name,
        "fixable": issue.fixable,
    }


def _classify_row_status(issues: list[RowIssue]) -> RowStatus:
    if any(issue.severity == "error" for issue in issues):
        return "error"
    if issues:
        return "incomplete"
    return "importable"


def _get_default_value(options: EmployeeImportOptions, field_name: str) -> str | None:
    raw_value = options.default_values.get(field_name)
    if field_name in {"full_name", "department_name", "job_profile"}:
        return _smart_title_case(raw_value)
    if field_name == "email":
        return _normalize_email(_auto_fix_email(raw_value) if options.auto_fix_enabled else raw_value)
    if field_name in {"employee_identifier", "employee_code"}:
        return _normalize_identifier(raw_value)
    if field_name == "status":
        return _normalize_status(raw_value)
    return _clean_text(raw_value)


def _get_row_override_map(options: EmployeeImportOptions) -> dict[int, dict[str, Any]]:
    row_override_map: dict[int, dict[str, Any]] = {}

    for row_override in options.row_overrides:
        payload = row_override.model_dump(exclude_unset=True)
        row_number = int(payload.pop("row_number"))
        row_override_map[row_number] = {
            key: value
            for key, value in payload.items()
            if key in EDITABLE_FIELDS
        }

    return row_override_map


def _normalize_employee_row(
    row_number: int,
    raw_row: dict[str, str],
    field_to_header: dict[str, str],
    row_override_map: dict[int, dict[str, Any]],
    options: EmployeeImportOptions,
) -> ParsedEmployeeRow:
    row_overrides = row_override_map.get(row_number, {})

    def resolve_value(field_name: str) -> str | None:
        if field_name in row_overrides:
            return _clean_text(row_overrides.get(field_name))
        source_header = field_to_header.get(field_name)
        if not source_header:
            return None
        return _clean_text(raw_row.get(source_header))

    raw_full_name = resolve_value("full_name")
    raw_email = resolve_value("email")
    raw_identifier = resolve_value("employee_identifier")
    raw_employee_code = resolve_value("employee_code")
    raw_department_name = resolve_value("department_name")
    raw_job_profile = resolve_value("job_profile")
    raw_status = resolve_value("status")

    full_name = _smart_title_case(raw_full_name) if options.auto_fix_enabled else _clean_text(raw_full_name)
    email_candidate = _auto_fix_email(raw_email) if options.auto_fix_enabled else raw_email
    email = _normalize_email(email_candidate)
    employee_identifier = _normalize_identifier(raw_identifier)
    employee_code = _normalize_identifier(raw_employee_code)
    department_name = (
        _smart_title_case(raw_department_name) if options.auto_fix_enabled else _clean_text(raw_department_name)
    )
    job_profile = (
        _smart_title_case(raw_job_profile) if options.auto_fix_enabled else _clean_text(raw_job_profile)
    )
    row_status = _normalize_status(raw_status)

    if email is None and raw_email is None and raw_identifier:
        inferred_email_candidate = _auto_fix_email(raw_identifier) if options.auto_fix_enabled else raw_identifier
        inferred_email = _normalize_email(inferred_email_candidate)
        if inferred_email is not None:
            email = inferred_email
            employee_identifier = None

    if department_name is None:
        department_name = _get_default_value(options, "department_name")
    if job_profile is None:
        job_profile = _get_default_value(options, "job_profile")
    if raw_status is None:
        row_status = _get_default_value(options, "status") or row_status

    issues: list[RowIssue] = []
    if full_name is None:
        issues.append(
            RowIssue(
                code="missing_full_name",
                severity="error",
                message="Nom complet manquant.",
                field_name="full_name",
                fixable=True,
            )
        )

    if raw_email is not None and email is None:
        issues.append(
            RowIssue(
                code="invalid_email",
                severity="error",
                message="Email invalide.",
                field_name="email",
                fixable=True,
            )
        )

    if email is None and employee_identifier is None and employee_code is None:
        issues.append(
            RowIssue(
                code="missing_identity",
                severity="error",
                message="Email, identifiant ou matricule requis.",
                field_name="email",
                fixable=True,
            )
        )

    if department_name is None:
        issues.append(
            RowIssue(
                code="missing_department",
                severity="warning",
                message="Departement manquant.",
                field_name="department_name",
                fixable=True,
            )
        )

    if job_profile is None:
        issues.append(
            RowIssue(
                code="missing_job_profile",
                severity="warning",
                message="Fonction manquante.",
                field_name="job_profile",
                fixable=True,
            )
        )

    if email is None and (employee_identifier is not None or employee_code is not None):
        issues.append(
            RowIssue(
                code="missing_email",
                severity="warning",
                message="Email non fourni: l'import restera possible avec l'identifiant ou le matricule.",
                field_name="email",
                fixable=True,
            )
        )

    identity_key = _build_identity_key(email, employee_identifier, employee_code)
    computed_row_status = _classify_row_status(issues)

    return ParsedEmployeeRow(
        row_number=row_number,
        full_name=full_name,
        email=email,
        employee_identifier=employee_identifier,
        employee_code=employee_code,
        department_name=department_name,
        job_profile=job_profile,
        status=row_status,
        identity_key=identity_key,
        row_status=computed_row_status,
        issues=issues,
    )


def _calculate_quality_score(rows: list[ParsedEmployeeRow], duplicate_rows: int, error_rows: int) -> int:
    if not rows:
        return 0

    completeness_points = 0
    possible_points = len(rows) * 5

    for row in rows:
        completeness_points += int(row.full_name is not None)
        completeness_points += int(row.identity_key is not None)
        completeness_points += int(row.email is not None)
        completeness_points += int(row.department_name is not None)
        completeness_points += int(row.job_profile is not None)

    completeness_ratio = completeness_points / possible_points
    validity_ratio = (len(rows) - error_rows) / len(rows)
    coherence_ratio = max(0.0, 1 - (duplicate_rows / len(rows)))

    score = round((completeness_ratio * 0.4 + validity_ratio * 0.35 + coherence_ratio * 0.25) * 100)
    return max(0, min(100, score))


def _build_suggestions(
    rows: list[ParsedEmployeeRow],
    field_mappings: list[dict[str, Any]],
    options: EmployeeImportOptions,
) -> list[dict[str, Any]]:
    suggestions: list[dict[str, Any]] = []

    def count_rows_with_issue(code: str) -> int:
        return sum(1 for row in rows if any(issue.code == code for issue in row.issues))

    missing_department_rows = count_rows_with_issue("missing_department")
    missing_job_profile_rows = count_rows_with_issue("missing_job_profile")
    invalid_email_rows = count_rows_with_issue("invalid_email")

    department_counter = Counter(
        row.department_name
        for row in rows
        if row.department_name is not None
    )
    suggested_department = department_counter.most_common(1)[0][0] if department_counter else None

    if missing_department_rows > 0:
        action_type = "apply_default_value" if suggested_department else "review_mapping"
        suggestions.append(
            {
                "id": "department-default",
                "title": f"Le departement n'est pas fourni pour {missing_department_rows} collaborateur(s).",
                "description": (
                    f"Souhaitez-vous attribuer '{suggested_department}' par defaut pour accelerer l'import ?"
                    if suggested_department
                    else "Associez une colonne ou appliquez une valeur par defaut pour harmoniser le fichier."
                ),
                "action_label": (
                    f"Attribuer {suggested_department}"
                    if suggested_department
                    else "Verifier le mapping"
                ),
                "action_type": action_type,
                "target_field": "department_name",
                "suggested_value": suggested_department,
                "affected_rows": missing_department_rows,
            }
        )

    if missing_job_profile_rows > 0:
        suggestions.append(
            {
                "id": "job-profile-followup",
                "title": f"Fonction manquante pour {missing_job_profile_rows} employe(s).",
                "description": "Vous pouvez completer ces informations apres import sans bloquer le chargement.",
                "action_label": "Completer apres import",
                "action_type": "complete_after_import",
                "target_field": "job_profile",
                "suggested_value": None,
                "affected_rows": missing_job_profile_rows,
            }
        )

    if invalid_email_rows > 0 and not options.auto_fix_enabled:
        suggestions.append(
            {
                "id": "auto-fix",
                "title": f"{invalid_email_rows} email(s) necessitent une correction.",
                "description": "La correction automatique applique les normalisations simples sans bloquer l'utilisateur.",
                "action_label": "Corriger automatiquement",
                "action_type": "auto_fix",
                "target_field": "email",
                "suggested_value": None,
                "affected_rows": invalid_email_rows,
            }
        )

    for field_mapping in field_mappings:
        if field_mapping["source_column"] is not None:
            continue
        if field_mapping["field_name"] not in {"full_name", "email", "department_name", "job_profile"}:
            continue
        if not field_mapping["suggested_columns"]:
            continue
        suggestions.append(
            {
                "id": f"map-{field_mapping['field_name']}",
                "title": f"Associer la colonne {field_mapping['label'].lower()}",
                "description": "Une colonne proche a ete detectee. Validez le mapping pour fiabiliser l'import.",
                "action_label": "Verifier le mapping",
                "action_type": "review_mapping",
                "target_field": field_mapping["field_name"],
                "suggested_value": field_mapping["suggested_columns"][0],
                "affected_rows": 0,
            }
        )

    return suggestions[:4]


def _analyze_employee_file(
    db: Session,
    *,
    file_name: str,
    content: bytes,
    options: EmployeeImportOptions | None = None,
) -> EmployeeFileAnalysis:
    _ensure_imported_employee_storage(db)
    resolved_options = options or EmployeeImportOptions()
    detected_format, headers, rows = _read_uploaded_rows(file_name, content)

    field_to_header, field_mappings, recognized_columns, missing_required_fields, warnings = _build_field_mappings(
        headers,
        resolved_options,
    )

    if not rows:
        warnings = [*warnings, "Aucune ligne exploitable n'a ete detectee dans le fichier."]

    existing_rows = db.execute(
        select(
            ImportedEmployee.email,
            ImportedEmployee.employee_identifier,
            ImportedEmployee.employee_code,
        )
    ).all()
    existing_emails = {
        normalized
        for email, _, _ in existing_rows
        if (normalized := _normalize_email(email)) is not None
    }
    existing_identifiers = {
        normalized
        for _, employee_identifier, _ in existing_rows
        if (normalized := _normalize_identifier(employee_identifier)) is not None
    }
    existing_codes = {
        normalized
        for _, _, employee_code in existing_rows
        if (normalized := _normalize_identifier(employee_code)) is not None
    }

    row_override_map = _get_row_override_map(resolved_options)

    seen_emails: set[str] = set()
    seen_identifiers: set[str] = set()
    seen_codes: set[str] = set()

    preview_rows: list[dict[str, Any]] = []
    importable_rows: list[ParsedEmployeeRow] = []
    parsed_rows: list[ParsedEmployeeRow] = []

    ready_rows = 0
    incomplete_rows = 0
    invalid_rows = 0
    duplicate_rows = 0

    for index, raw_row in enumerate(rows, start=2):
        parsed_row = _normalize_employee_row(
            index,
            raw_row,
            field_to_header,
            row_override_map,
            resolved_options,
        )

        has_blocking_error = any(issue.severity == "error" for issue in parsed_row.issues)
        duplicate_messages: list[str] = []

        if not has_blocking_error:
            if parsed_row.email:
                if parsed_row.email in seen_emails:
                    duplicate_messages.append("Doublon sur l'email present plusieurs fois dans le fichier.")
                elif parsed_row.email in existing_emails:
                    duplicate_messages.append("Doublon sur l'email deja enregistre dans l'application.")

            if parsed_row.employee_identifier:
                if parsed_row.employee_identifier in seen_identifiers:
                    duplicate_messages.append(
                        "Doublon sur l'identifiant present plusieurs fois dans le fichier."
                    )
                elif parsed_row.employee_identifier in existing_identifiers:
                    duplicate_messages.append(
                        "Doublon sur l'identifiant deja enregistre dans l'application."
                    )

            if parsed_row.employee_code:
                if parsed_row.employee_code in seen_codes:
                    duplicate_messages.append("Doublon sur le matricule present plusieurs fois dans le fichier.")
                elif parsed_row.employee_code in existing_codes:
                    duplicate_messages.append(
                        "Doublon sur le matricule deja enregistre dans l'application."
                    )

            if duplicate_messages:
                parsed_row.issues.extend(
                    [
                        RowIssue(
                            code="duplicate_record",
                            severity="error",
                            message=message,
                            field_name="email",
                            fixable=False,
                        )
                        for message in duplicate_messages
                    ]
                )
                parsed_row.duplicate_reason = " ".join(duplicate_messages)
                parsed_row.row_status = "error"
            else:
                if parsed_row.email:
                    seen_emails.add(parsed_row.email)
                if parsed_row.employee_identifier:
                    seen_identifiers.add(parsed_row.employee_identifier)
                if parsed_row.employee_code:
                    seen_codes.add(parsed_row.employee_code)

        parsed_row.row_status = _classify_row_status(parsed_row.issues)
        parsed_rows.append(parsed_row)

        has_duplicate_issue = any(issue.code == "duplicate_record" for issue in parsed_row.issues)
        has_non_duplicate_error = any(
            issue.severity == "error" and issue.code != "duplicate_record"
            for issue in parsed_row.issues
        )

        if parsed_row.row_status == "error":
            if has_non_duplicate_error:
                invalid_rows += 1
            elif has_duplicate_issue:
                duplicate_rows += 1
        elif parsed_row.row_status == "incomplete":
            incomplete_rows += 1
            importable_rows.append(parsed_row)
        else:
            ready_rows += 1
            importable_rows.append(parsed_row)

        if len(preview_rows) < PREVIEW_ROW_LIMIT:
            preview_rows.append(
                {
                    "row_number": parsed_row.row_number,
                    "full_name": parsed_row.full_name,
                    "email": parsed_row.email,
                    "employee_identifier": parsed_row.employee_identifier,
                    "employee_code": parsed_row.employee_code,
                    "department_name": parsed_row.department_name,
                    "job_profile": parsed_row.job_profile,
                    "status": parsed_row.status,
                    "row_status": parsed_row.row_status,
                    "issues": [_serialize_issue(issue) for issue in parsed_row.issues],
                    "errors": [
                        issue.message
                        for issue in parsed_row.issues
                        if issue.severity == "error"
                    ],
                    "duplicate_reason": parsed_row.duplicate_reason,
                }
            )

    error_rows = invalid_rows + duplicate_rows
    valid_rows = ready_rows + incomplete_rows
    anomalies_count = incomplete_rows + error_rows
    fixable_anomalies = sum(
        1
        for row in parsed_rows
        for issue in row.issues
        if issue.fixable
    )
    quality_score = _calculate_quality_score(parsed_rows, duplicate_rows, error_rows)
    suggestions = _build_suggestions(parsed_rows, field_mappings, resolved_options)

    global_notice = None
    if anomalies_count > 0:
        global_notice = f"{anomalies_count} anomalie(s) detectee(s) - correction recommandee."
    elif rows:
        global_notice = "Aucune anomalie bloquante detectee. Le fichier est pret a etre importe."

    return EmployeeFileAnalysis(
        file_name=file_name,
        detected_format=detected_format,
        total_rows=len(rows),
        valid_rows=valid_rows,
        ready_rows=ready_rows,
        incomplete_rows=incomplete_rows,
        invalid_rows=invalid_rows,
        duplicate_rows=duplicate_rows,
        error_rows=error_rows,
        quality_score=quality_score,
        anomalies_count=anomalies_count,
        fixable_anomalies=fixable_anomalies,
        global_notice=global_notice,
        recognized_columns=recognized_columns,
        available_columns=headers,
        field_mappings=field_mappings,
        missing_required_fields=missing_required_fields,
        warnings=warnings,
        suggestions=suggestions,
        preview_rows=preview_rows,
        importable_rows=importable_rows,
    )


def preview_employee_import(
    db: Session,
    *,
    file_name: str,
    content: bytes,
    options: EmployeeImportOptions | None = None,
) -> dict[str, Any]:
    analysis = _analyze_employee_file(db, file_name=file_name, content=content, options=options)
    return {
        "file_name": analysis.file_name,
        "detected_format": analysis.detected_format,
        "total_rows": analysis.total_rows,
        "valid_rows": analysis.valid_rows,
        "ready_rows": analysis.ready_rows,
        "incomplete_rows": analysis.incomplete_rows,
        "invalid_rows": analysis.invalid_rows,
        "duplicate_rows": analysis.duplicate_rows,
        "error_rows": analysis.error_rows,
        "quality_score": analysis.quality_score,
        "anomalies_count": analysis.anomalies_count,
        "fixable_anomalies": analysis.fixable_anomalies,
        "global_notice": analysis.global_notice,
        "recognized_columns": analysis.recognized_columns,
        "available_columns": analysis.available_columns,
        "field_mappings": analysis.field_mappings,
        "missing_required_fields": analysis.missing_required_fields,
        "warnings": analysis.warnings,
        "suggestions": analysis.suggestions,
        "preview_rows": analysis.preview_rows,
    }


def import_employees(
    db: Session,
    *,
    file_name: str,
    content: bytes,
    options: EmployeeImportOptions | None = None,
) -> dict[str, Any]:
    analysis = _analyze_employee_file(db, file_name=file_name, content=content, options=options)

    if not analysis.importable_rows:
        return {
            "file_name": analysis.file_name,
            "detected_format": analysis.detected_format,
            "total_rows": analysis.total_rows,
            "imported_count": 0,
            "incomplete_count": 0,
            "skipped_count": analysis.error_rows,
            "duplicate_count": analysis.duplicate_rows,
            "invalid_count": analysis.invalid_rows,
            "rejected_count": analysis.error_rows,
            "quality_score": analysis.quality_score,
            "recognized_columns": analysis.recognized_columns,
            "warnings": [
                *analysis.warnings,
                "Aucun employe supplementaire n'a ete enregistre.",
            ],
        }

    for row in analysis.importable_rows:
        db.add(
            ImportedEmployee(
                full_name=row.full_name or "Collaborateur sans nom",
                identity_key=row.identity_key or "",
                email=row.email,
                employee_identifier=row.employee_identifier,
                employee_code=row.employee_code,
                department_name=row.department_name,
                job_profile=row.job_profile,
                status=row.status,
                source_filename=file_name,
                source_format=analysis.detected_format,
            )
        )

    db.commit()

    return {
        "file_name": analysis.file_name,
        "detected_format": analysis.detected_format,
        "total_rows": analysis.total_rows,
        "imported_count": analysis.valid_rows,
        "incomplete_count": analysis.incomplete_rows,
        "skipped_count": analysis.error_rows,
        "duplicate_count": analysis.duplicate_rows,
        "invalid_count": analysis.invalid_rows,
        "rejected_count": analysis.error_rows,
        "quality_score": analysis.quality_score,
        "recognized_columns": analysis.recognized_columns,
        "warnings": analysis.warnings,
    }


def list_imported_employees(
    db: Session,
    *,
    offset: int = 0,
    limit: int = 50,
    search: str | None = None,
    status_filter: str | None = None,
) -> dict[str, Any]:
    _ensure_imported_employee_storage(db)
    statement = select(ImportedEmployee)

    normalized_search = _normalize_search(search)
    if normalized_search:
        like_pattern = f"%{normalized_search}%"
        statement = statement.where(
            or_(
                func.lower(ImportedEmployee.full_name).like(like_pattern),
                func.lower(func.coalesce(ImportedEmployee.email, "")).like(like_pattern),
                func.lower(func.coalesce(ImportedEmployee.employee_identifier, "")).like(like_pattern),
                func.lower(func.coalesce(ImportedEmployee.employee_code, "")).like(like_pattern),
                func.lower(func.coalesce(ImportedEmployee.department_name, "")).like(like_pattern),
            )
        )

    normalized_status = _normalize_search(status_filter)
    if normalized_status in {"active", "inactive", "suspended"}:
        statement = statement.where(ImportedEmployee.status == normalized_status)

    total = db.scalar(select(func.count()).select_from(statement.subquery())) or 0
    items = list(
        db.scalars(
            statement
            .order_by(
                ImportedEmployee.full_name.asc(),
                ImportedEmployee.id.asc(),
            )
            .offset(offset)
            .limit(limit)
        )
    )

    return {
        "total": int(total),
        "offset": offset,
        "limit": limit,
        "items": items,
    }
